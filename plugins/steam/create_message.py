import os
import json
import pathlib

import pandas as pd
import openpyxl
from airflow import AirflowException
from airflow.models import BaseOperator
from airflow.utils.context import Context
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class CreateMessage(BaseOperator):
    """
    Формирует файлы с сообщениями
    и таблицы excel
    """

    def __init__(
        self,
        config: dict,
        path_save: str,
        path_for_json: str,
        date_period_type: str,
        columns_for_table: list,
        columns_increment_for_table: list,
        columns_increment_for_table_csv: list,
        domain_code: str,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.config = config
        self.path_save = path_save
        self.path_for_json = path_for_json
        self.date_period_type = date_period_type
        self.columns_for_table = columns_for_table
        self.columns_increment_for_table = columns_increment_for_table
        self.columns_increment_for_table_csv = columns_increment_for_table_csv
        self.domain_code = domain_code

    def _check_args(self):
        if self.date_period_type not in ["day", "week", "month", "year"]:
            raise AirflowException("Неверный тип аргумента date_period_type")

    def _find_actual_date(self, context: Context) -> str:
        actual_date = context["data_interval_start"].start_of(self.date_period_type)
        return (
            actual_date.day_of_week,
            actual_date.format("YYYY-MM-DD"),
            context["data_interval_start"].format("YYYY-MM-DD"),
        )

    def _text_data(self, data: dict, order_columns: list):
        message_data = ""
        for i, _ in enumerate(data["game_name"]):
            data_list = [str(data[key][i]) for key in order_columns]
            message_data += " | ".join(data_list)
            message_data += "\n"
        return message_data

    def _formatt_excel(self, path_for_file: str) -> None:
        workbook = openpyxl.load_workbook(path_for_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.auto_filter.ref = sheet.dimensions
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[cell.column_letter].width = adjusted_width

            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = PatternFill(
                        start_color="55b1ec", end_color="55b1ec", fill_type="solid"
                    )

        workbook.save(path_for_file)

    def _create_xlsx_increment(
        self, data: dict, columns_csv: list, columns_domen: list, path_save_data: str
    ):
        print("columns_csv", columns_csv)
        print("columns_domen", columns_domen)
        print("data", data.keys())
        df = pd.DataFrame(data)[columns_csv]
        df.columns = [columns_domen[i] for i in columns_csv]
        print(df)
        df.to_excel(path_save_data, index=False)

        print("path_save_data", path_save_data)

        self._formatt_excel(path_save_data)

    def _create_save_message(
        self,
        data,
        actual_date: str,
        day_of_week: int,
        config: dict,
        path_save: str,
    ):
        col_dict = " | ".join(
            [config["columns_name"][key] for key in self.columns_for_table]
        )
        col_dict_increment = " | ".join(
            [config["columns_name"][key] for key in self.columns_increment_for_table]
        )

        actual_text_data = self._text_data(
            data=data["actual_data"], order_columns=self.columns_for_table
        )

        message_top_now = Template(config["message_top_now"]).render(
            ds=actual_date,
            day_of_week=config["day_of_week_translate"][day_of_week],
            date_period_type=config["date_period_type_translate"][
                self.date_period_type
            ],
            col=col_dict,
            pages_data=actual_text_data,
        )

        new_in_top_text_data = self._text_data(
            data=data["new_in_top"], order_columns=self.columns_for_table
        )
        message_new_games = Template(config["message_new_games"]).render(
            date_period_type=config["date_period_type_translate"][
                self.date_period_type
            ],
            col=col_dict,
            pages_data=new_in_top_text_data,
        )

        go_out_from_top_text_data = self._text_data(
            data=data["go_out_from_top"], order_columns=self.columns_for_table
        )
        message_go_out_pages = Template(config["message_go_out_games"]).render(
            col=col_dict,
            pages_data=go_out_from_top_text_data,
        )

        stay_in_top_text_data = self._text_data(
            data=data["stay_in_top"], order_columns=self.columns_increment_for_table
        )

        message_difference_games = Template(config["message_difference_games"]).render(
            col=col_dict_increment,
            pages_data=stay_in_top_text_data,
        )

        for message, message_type in zip(
            [
                message_top_now,
                message_new_games,
                message_go_out_pages,
                message_difference_games,
            ],
            ["TopNow", "NewInTop", "GoOut", "Diff"],
        ):
            file_name = f"{self.date_period_type}_{message_type}_{actual_date}.txt"
            file_save_path = os.path.join(path_save, file_name)

            with open(file_save_path, "w") as f:
                f.write(message)

        ### Сохраняем данные как xlsx

        path_save_xlsx = os.path.join(path_save, "xlsx")
        pathlib.Path(path_save_xlsx).mkdir(parents=True, exist_ok=True)

        for data_type, message_type, columns_csv in zip(
            [
                "actual_data",
                "new_in_top",
                "go_out_from_top",
                "stay_in_top",
            ],
            ["TopNow", "NewInTop", "GoOut", "Diff"],
            [
                self.columns_for_table,
                self.columns_for_table,
                self.columns_for_table,
                self.columns_increment_for_table_csv,
            ],
        ):
            file_name = f"{self.date_period_type}_{message_type}_{actual_date}.xlsx"
            file_save_path = os.path.join(path_save_xlsx, file_name)

            print(message_type)

            self._create_xlsx_increment(
                data=data[data_type],
                columns_csv=columns_csv,
                columns_domen=config["columns_name"],
                path_save_data=file_save_path,
            )

        file_excel_name = f"{self.date_period_type}_data_{actual_date}.xlsx"

        file_save_path = os.path.join(path_save_xlsx, file_excel_name)

        excel_files = [
            i
            for i in os.listdir(path_save_xlsx)
            if i.endswith(".xlsx")
            and i.startswith(self.date_period_type)
            and i != file_excel_name
            # нужно проверять на имя вновь создаваемого файла,
            # потому что алгорит может быть на пройд. интервалах
        ]

        pd.DataFrame().to_excel(file_save_path)

        combined_wb = load_workbook(file_save_path)
        sheet = combined_wb["Sheet1"]
        combined_wb.remove(sheet)

        for file in excel_files:
            file_path = os.path.join(path_save_xlsx, file)
            wb = load_workbook(file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                combined_wb.create_sheet(title=file)
                combined_sheet = combined_wb[file]
                for row in sheet.iter_rows():
                    combined_sheet.append([cell.value for cell in row])
        combined_wb.save(file_save_path)
        self._formatt_excel(file_save_path)

    def execute(self, context: Context) -> None:
        self._check_args()

        day_of_week, actual_date, ds = self._find_actual_date(context)
        path_save = os.path.join(self.path_save, self.domain_code, ds)
        pathlib.Path(path_save).mkdir(parents=True, exist_ok=True)

        with open(
            os.path.join(
                self.path_for_json,
                self.date_period_type,
                f"{actual_date}_{self.date_period_type}.json",
            )
        ) as f:
            data = json.load(f)

        self._create_save_message(
            data=data,
            actual_date=actual_date,
            day_of_week=day_of_week,
            config=self.config,
            path_save=path_save,
        )
